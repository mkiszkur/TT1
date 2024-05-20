import os
import pandas as pd
import openpyxl
import xlrd
import logging

# Configura el logging para capturar errores
def setup_logging(log_path):
    logging.basicConfig(filename=log_path, filemode='w', level=logging.ERROR)

# Función para procesar archivos Excel y agregar datos al CSV usando openpyxl
def add_data_to_csv_xlsx(excel_file, csv_file):
    try:
        # Carga la solapa específica
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook['Metadata - Indicators']
        data = {
            'INDICATOR_CODE': sheet['A2'].value,
            'FILE_NAME': excel_file,
            'INDICATOR_NAME': sheet['B2'].value,
            'SOURCE_NOTE': sheet['C2'].value,
            'SOURCE_ORGANIZATION': sheet['D2'].value
        }
        df = pd.DataFrame([data])
        df.to_csv(csv_file, mode='a', header=False, index=False)
    except Exception as e:
        logging.error(f'Error processing file {excel_file}: {e}')

# Función para procesar archivos Excel y agregar datos al CSV usando xlrd
def add_data_to_csv_xls(excel_file, csv_file):
    try:
        # Abre el archivo xls
        workbook = xlrd.open_workbook(excel_file)
        sheet = workbook.sheet_by_name('Metadata - Indicators')
        data = {
            'INDICATOR_CODE': sheet.cell(1, 0).value,
            'FILE_NAME': excel_file,
            'INDICATOR_NAME': sheet.cell(1, 1).value,
            'SOURCE_NOTE': sheet.cell(1, 2).value,
            'SOURCE_ORGANIZATION': sheet.cell(1, 3).value
        }
        print (data)
        df = pd.DataFrame([data])
        df.to_csv(csv_file, mode='a', header=False, index=False)
    except Exception as e:
        logging.error(f'Error processing file {excel_file}: {e}')

# Función principal para iterar sobre los archivos Excel en un directorio
def process_folder(directory, csv_file):
    log_path = csv_file.replace('.csv', '.log')
    setup_logging(log_path)
    
    for file in os.listdir(directory):
        print (file)
        if file.endswith('.xlsx'):
            excel_file = os.path.join(directory, file)
            add_data_to_csv_xlsx(excel_file, csv_file)
        elif file.endswith('.xls'):
            excel_file = os.path.join(directory, file)
            add_data_to_csv_xls(excel_file, csv_file)
        #print (excel_file)
# Reemplaza con el directorio de tus archivos y el nombre de tu archivo CSV

directory_path = '/Users/miguelkiszkurno/Documents/TT1/data/'
csv_file_path = '/Users/miguelkiszkurno/Documents/TT1/indicadores.csv'

# Inicia el procesamiento
process_folder(directory_path, csv_file_path)